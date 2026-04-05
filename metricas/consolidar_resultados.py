"""
consolidar_resultados.py
Lee los 3 JSONs de resultados (TotalSegmentator, MedSAM, YOLO) y rellena
automáticamente la tabla comparativa Excel generada previamente.

Uso:
    python consolidar_resultados.py \
        --ts  resultados_totalsegmentator.json \
        --ms  resultados_medsam.json \
        --yo  resultados_yolo.json \
        --xlsx comparativa_modelos_segmentacion.xlsx \
        --out  comparativa_completa.xlsx
"""

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


COLOR_GOOD    = "C6EFCE"  # verde claro
COLOR_WARN    = "FFEB9C"  # amarillo
COLOR_BAD     = "FFC7CE"  # rojo claro
COLOR_NEUTRAL = "FAFAFA"

def color_for_dice(d):
    if d is None: return COLOR_NEUTRAL
    if d >= 0.80: return COLOR_GOOD
    if d >= 0.50: return COLOR_WARN
    return COLOR_BAD

def fmt(v, decimals=4):
    if v is None: return "—"
    if isinstance(v, float): return round(v, decimals)
    return v

def load_json(path):
    if path is None or not Path(path).exists():
        return None
    with open(path, encoding="utf-8") as f:
        return json.load(f)

def build_structure_lookup(report):
    """Devuelve dict estructura_name -> entry"""
    if report is None:
        return {}
    return {e["estructura"]: e for e in report.get("detalle_por_estructura", [])}


def fill_excel(ts_report, ms_report, yo_report, template_path, out_path):
    wb = load_workbook(template_path)
    ws = wb.active

    # ── Hoja de resumen ejecutivo (nueva hoja) ─────────────────────────────────
    if "Resumen" in wb.sheetnames:
        del wb["Resumen"]
    ws_res = wb.create_sheet("Resumen", 0)

    models = [
        ("TotalSegmentator\n(nnU-Net)", ts_report),
        ("MedSAM",                      ms_report),
        ("YOLO\n(YOLOv8x-seg)",         yo_report),
    ]

    # Encabezado resumen
    ws_res.column_dimensions['A'].width = 38
    ws_res.column_dimensions['B'].width = 22
    ws_res.column_dimensions['C'].width = 22
    ws_res.column_dimensions['D'].width = 22

    def hdr(cell, text, bg="1F4E79", fg="FFFFFF", bold=True, size=10):
        cell.value = text
        cell.font = Font(bold=bold, color=fg, size=size, name="Arial")
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def dat(cell, text, bg=COLOR_NEUTRAL, bold=False, align="center"):
        cell.value = text
        cell.font = Font(bold=bold, size=10, name="Arial", color="222222")
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

    ws_res.row_dimensions[1].height = 36
    ws_res.merge_cells("A1:D1")
    hdr(ws_res["A1"], "Tabla Comparativa — Resumen Ejecutivo de Modelos de Segmentación", bg="0D2137", size=12)

    ws_res.row_dimensions[2].height = 35
    for col, (name, _) in zip(['B','C','D'], models):
        hdr(ws_res[f"{col}2"], name, bg="1F4E79")
    hdr(ws_res["A2"], "Métrica", bg="1F4E79")

    metrics_summary = [
        ("Dice promedio (vs GT)",       lambda r: fmt(r["metricas_agregadas"]["dice_promedio"], 3) if r else "N/D"),
        ("Dice mediana",                lambda r: fmt(r["metricas_agregadas"]["dice_mediana"],  3) if r else "N/D"),
        ("IoU promedio",                lambda r: fmt(r["metricas_agregadas"]["iou_promedio"],  3) if r else "N/D"),
        ("Dice mínimo",                 lambda r: fmt(r["metricas_agregadas"]["dice_min"],      3) if r else "N/D"),
        ("Dice máximo",                 lambda r: fmt(r["metricas_agregadas"]["dice_max"],      3) if r else "N/D"),
        ("Estructuras comparadas",      lambda r: f"{r['estructuras']['comparadas_exitosamente']} / {r['estructuras']['total_evaluadas']}" if r else "N/D"),
        ("Tiempo total (minutos)",      lambda r: fmt(r["tiempo_inferencia"]["total_minutos"], 1) if r else "N/D"),
        ("Dispositivo",                 lambda r: r.get("dispositivo", "N/D") if r else "N/D"),
        ("Nota metodológica",           lambda r: r.get("nota_metodologica", "—") if r else "N/D"),
    ]

    for i, (label, fn) in enumerate(metrics_summary, start=3):
        ws_res.row_dimensions[i].height = 40 if i == i+8 else 28
        dat(ws_res[f"A{i}"], label, align="left")
        for col, (_, rpt) in zip(['B','C','D'], models):
            val = fn(rpt)
            # Color por Dice
            bg = COLOR_NEUTRAL
            if "Dice" in label and isinstance(val, float):
                bg = color_for_dice(val)
            dat(ws_res[f"{col}{i}"], val, bg=bg)

    # ── Hoja de detalle por estructura ────────────────────────────────────────
    if "Detalle Estructuras" in wb.sheetnames:
        del wb["Detalle Estructuras"]
    ws_det = wb.create_sheet("Detalle Estructuras", 1)

    ws_det.column_dimensions['A'].width = 30
    ws_det.column_dimensions['B'].width = 14
    ws_det.column_dimensions['C'].width = 14
    ws_det.column_dimensions['D'].width = 14
    ws_det.column_dimensions['E'].width = 14
    ws_det.column_dimensions['F'].width = 14
    ws_det.column_dimensions['G'].width = 14

    ws_det.row_dimensions[1].height = 30
    ws_det.merge_cells("A1:G1")
    hdr(ws_det["A1"], "Detalle por Estructura Anatómica — Dice Score y Volúmenes", bg="0D2137", size=11)

    headers_det = [
        "Estructura",
        "TS Dice", "TS Vol pred (ml)",
        "MedSAM Dice", "MedSAM Vol pred (ml)",
        "YOLO Dice", "YOLO Vol pred (ml)",
    ]
    ws_det.row_dimensions[2].height = 32
    for col_idx, h in enumerate(headers_det, start=1):
        from openpyxl.utils import get_column_letter
        cell = ws_det.cell(row=2, column=col_idx)
        hdr(cell, h, bg="2E75B6", size=9)

    ts_lu = build_structure_lookup(ts_report)
    ms_lu = build_structure_lookup(ms_report)
    yo_lu = build_structure_lookup(yo_report)

    all_structures = sorted(set(
        list(ts_lu.keys()) + list(ms_lu.keys()) + list(yo_lu.keys())
    ))

    for row_idx, struct in enumerate(all_structures, start=3):
        ws_det.row_dimensions[row_idx].height = 22
        dat(ws_det.cell(row=row_idx, column=1), struct, align="left")

        for col_offset, lu in [(2, ts_lu), (4, ms_lu), (6, yo_lu)]:
            entry = lu.get(struct)
            dice_val = entry["dice"] if entry else None
            vol_val  = entry.get("volumen_pred_ml") if entry else None
            bg = color_for_dice(dice_val)
            dat(ws_det.cell(row=row_idx, column=col_offset),
                fmt(dice_val, 3) if dice_val is not None else "—", bg=bg)
            dat(ws_det.cell(row=row_idx, column=col_offset+1),
                fmt(vol_val, 1) if vol_val is not None else "—")

    # ── Rellenar tabla comparativa original ───────────────────────────────────
    # Mapeo de celdas de la tabla original → valores de los reportes
    # La tabla original tiene columnas C=TS, D=MedSAM, E=YOLO
    # Aquí escribimos las métricas clave en las filas correspondientes

    ws_orig = wb["Comparativa Modelos"]

    def find_row_by_label(ws, label_substr: str, col='A') -> int | None:
        for row in ws.iter_rows():
            for cell in row:
                if cell.column_letter == col and cell.value and label_substr.lower() in str(cell.value).lower():
                    return cell.row
        return None

    def write_metric(ws, row, col_c_val, col_d_val, col_e_val, bg_fn=None):
        for col, val in zip(['C','D','E'], [col_c_val, col_d_val, col_e_val]):
            cell = ws[f"{col}{row}"]
            cell.value = val
            cell.font = Font(size=10, name="Arial", color="111111")
            bg = bg_fn(val) if (bg_fn and isinstance(val, float)) else COLOR_NEUTRAL
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Helper para extraer valores
    def get_agg(report, key):
        if report is None: return "N/D"
        v = report["metricas_agregadas"].get(key)
        return round(v, 3) if v is not None else "N/D"

    def get_time(report, key="total_minutos"):
        if report is None: return "N/D"
        v = report["tiempo_inferencia"].get(key)
        return f"{round(v, 1)} min" if v is not None else "N/D"

    def get_struct_count(report):
        if report is None: return "N/D"
        e = report["estructuras"]
        return f"{e['comparadas_exitosamente']} / {e['total_evaluadas']}"

    # Rellenar filas clave de la hoja original
    fill_map = {
        "N° estructuras detectadas vs. segmentación manual": (
            get_struct_count(ts_report), get_struct_count(ms_report), get_struct_count(yo_report)
        ),
        "Tiempo por caso (GPU)": (
            get_time(ts_report), get_time(ms_report), get_time(yo_report)
        ),
        "Revisión radiológica": (
            "Pendiente", "Pendiente", "Pendiente"
        ),
        "Recomendación final": (
            "Apto ✓" if ts_report else "N/D",
            "Parcial ⚠" if ms_report else "N/D",
            "No apto ✗" if yo_report else "N/D",
        ),
        "Adecuación para segmentación volumétrica 3D": (
            "✓ Nativa 3D", "⚠ 2D slice-a-slice", "✗ No volumétrico"
        ),
        "Soporte multiestructura (una sola inferencia)": (
            "✓ Multi-task", "✗ Un prompt por estructura", "✗ Clase genérica"
        ),
        "Compatibilidad con NIfTI directo": (
            "✓", "✓ (vía nibabel)", "✓ (vía nibabel)"
        ),
    }

    for label, (ts_val, ms_val, yo_val) in fill_map.items():
        row_num = find_row_by_label(ws_orig, label, col='A')
        if row_num:
            write_metric(ws_orig, row_num, ts_val, ms_val, yo_val)

    wb.save(str(out_path))
    print(f"[ok] Excel consolidado guardado en: {out_path}")


def main():
    parser = argparse.ArgumentParser(description="Consolida resultados JSON en tabla Excel")
    parser.add_argument("--ts",   default=None,  help="JSON de TotalSegmentator")
    parser.add_argument("--ms",   default=None,  help="JSON de MedSAM")
    parser.add_argument("--yo",   default=None,  help="JSON de YOLO")
    parser.add_argument("--xlsx", required=True, help="Excel plantilla (tabla comparativa)")
    parser.add_argument("--out",  default="comparativa_completa.xlsx", help="Excel de salida")
    args = parser.parse_args()

    ts = load_json(args.ts)
    ms = load_json(args.ms)
    yo = load_json(args.yo)

    loaded = sum(1 for x in [ts, ms, yo] if x is not None)
    print(f"[info] Reportes cargados: {loaded}/3")
    if ts: print(f"  ✓ TotalSegmentator — Dice={ts['metricas_agregadas']['dice_promedio']}")
    if ms: print(f"  ✓ MedSAM           — Dice={ms['metricas_agregadas']['dice_promedio']}")
    if yo: print(f"  ✓ YOLO             — Dice={yo['metricas_agregadas']['dice_promedio']}")

    fill_excel(ts, ms, yo, args.xlsx, args.out)


if __name__ == "__main__":
    main()
